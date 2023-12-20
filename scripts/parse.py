import openpyxl
import json


XLSX_PATH = "resources/Categorized AIO Command List v9.3 - DARK.xlsx"


def slot_to_dir(slot):
    match slot:
        case "FACE":
            return "catalog/face/"
        case "face":
            return "catalog/face/"
        case "FEET":
            return "catalog/feet/"
        case "feet":
            return "catalog/feet/"
        case "HEAD":
            return "catalog/head/"
        case "head":
            return "catalog/head/"
        case "INNER TORSO":
            return "catalog/inner-torso/"
        case "inner_torso":
            return "catalog/inner-torso/"
        case "LEGS":
            return "catalog/legs/"
        case "legs":
            return "catalog/legs/"
        case "OUTER TORSO":
            return "catalog/outer-torso/"
        case "outer_torso":
            return "catalog/outer-torso/"
        case "OUTFIT":
            return "catalog/outfit/"
        case "outfit":
            return "catalog/outfit/"
        case _:
            print("Unknown slot: " + slot)
            return ""

def slot_to_image(slot, TweakDB_id, variant):
    if slot in ["FACE", "FEET", "HEAD", "INNER TORSO", "LEGS", "OUTER TORSO", "OUTFIT"]:
        return slot_to_dir(slot) + TweakDB_id + "_" + variant + ".png"
    
    print("Unknown slot: " + slot)
    return ""

def xlsx_to_catalog(xlsx):
    face = {}
    feet = {}
    head = {}
    inner_torso = {}
    legs = {}
    outer_torso = {}
    outfit = {}

    workbook = openpyxl.load_workbook(xlsx)
    clothes = workbook[workbook.sheetnames[-2]]

    for i in range(8, clothes.max_row + 1):
        slot = clothes.cell(i, 2).value
        tag = clothes.cell(i, 3).value
        command = clothes.cell(i, 4).value
        name_female = clothes.cell(i, 5).value
        name_male = clothes.cell(i, 6).value
        TweakDB_id = command.split("\"")[1].split("\"")[0]
        
        if tag == "SPAWN0 MODS":
            continue
        
        item = {
            "command": command,
            "name_male": name_male,
            "image_male": slot_to_image(slot, TweakDB_id, "male"),
            "name_female": name_female,
            "image_female": slot_to_image(slot, TweakDB_id, "female"),
        }
        
        match slot:
            case "FACE":
                face[TweakDB_id] = item
            case "FEET":
                feet[TweakDB_id] = item
            case "HEAD":
                head[TweakDB_id] = item
            case "INNER TORSO":
                inner_torso[TweakDB_id] = item
            case "LEGS":
                legs[TweakDB_id] = item
            case "OUTER TORSO":
                outer_torso[TweakDB_id] = item
            case "OUTFIT":
                outfit[TweakDB_id] = item
            case _:
                print("Unknown slot: " + slot)
    
    return {
        "face": face,
        "feet": feet,
        "head": head,
        "inner_torso": inner_torso,
        "legs": legs,
        "outer_torso": outer_torso,
        "outfit": outfit,
    }

def catalog_to_json(catalog):
    for slot in catalog:
        with open(slot_to_dir(slot) + "catalog.json", "w") as of:
            json.dump(catalog[slot], of, indent=4)


catalog = xlsx_to_catalog(XLSX_PATH)

# for slot in catalog:
#     print(slot + " " + str(len(catalog[slot])))

catalog_to_json(catalog)
